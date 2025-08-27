"""
エクスポート機能モジュール
CSVとExcel形式での結果エクスポートを担当
"""

import io
import csv
from typing import List, Dict, Any, Optional
import streamlit as st

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportManager:
    """エクスポート管理クラス"""
    
    def __init__(self):
        """ExportManagerの初期化"""
        self.supported_formats = ['txt', 'csv', 'xlsx']
    
    def export_to_csv(self, results: List[Dict[str, Any]], filename: str = "ocr_results.csv") -> Optional[bytes]:
        """
        結果をCSV形式でエクスポート
        
        Args:
            results: OCR結果のリスト
            filename: 出力ファイル名
            
        Returns:
            CSVデータのバイト列
        """
        if not results:
            return None
        
        try:
            # CSVデータの準備
            csv_data = []
            
            # ヘッダー行
            headers = ['画像番号', 'ファイル名', '抽出テキスト', '言語', '処理時間(秒)', 'ステータス', 'エラー']
            csv_data.append(headers)
            
            # データ行
            for i, result in enumerate(results):
                row = [
                    f"画像 {i+1}",
                    result.get('filename', f'image_{i+1}'),
                    result.get('text', ''),
                    result.get('language', 'unknown'),
                    f"{result.get('processing_time', 0):.2f}",
                    '成功' if result.get('success') else '失敗',
                    result.get('error', '')
                ]
                csv_data.append(row)
            
            # CSVファイルの作成
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_ALL, encoding='utf-8')
            writer.writerows(csv_data)
            
            return output.getvalue().encode('utf-8-sig')  # BOM付きUTF-8
            
        except Exception as e:
            st.error(f"CSVエクスポートに失敗しました: {str(e)}")
            return None
    
    def export_to_excel(self, results: List[Dict[str, Any]], filename: str = "ocr_results.xlsx") -> Optional[bytes]:
        """
        結果をExcel形式でエクスポート
        
        Args:
            results: OCR結果のリスト
            filename: 出力ファイル名
            
        Returns:
            Excelデータのバイト列
        """
        if not OPENPYXL_AVAILABLE:
            st.error("OpenPyXLライブラリがインストールされていません。Excel機能を使用するには 'pip install openpyxl' を実行してください。")
            return None
        
        if not results:
            return None
        
        try:
            # ワークブックの作成
            wb = Workbook()
            
            # メインシート（結果）
            ws_results = wb.active
            ws_results.title = "OCR結果"
            
            # ヘッダーの設定
            headers = ['画像番号', 'ファイル名', '抽出テキスト', '言語', '処理時間(秒)', 'ステータス', 'エラー']
            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # データの追加
            for i, result in enumerate(results, 2):
                ws_results.cell(row=i, column=1, value=f"画像 {i-1}")
                ws_results.cell(row=i, column=2, value=result.get('filename', f'image_{i-1}'))
                ws_results.cell(row=i, column=3, value=result.get('text', ''))
                ws_results.cell(row=i, column=4, value=result.get('language', 'unknown'))
                ws_results.cell(row=i, column=5, value=f"{result.get('processing_time', 0):.2f}")
                
                status_cell = ws_results.cell(row=i, column=6, value='成功' if result.get('success') else '失敗')
                if result.get('success'):
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                ws_results.cell(row=i, column=7, value=result.get('error', ''))
            
            # 列幅の自動調整
            for column in ws_results.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_results.column_dimensions[column_letter].width = adjusted_width
            
            # メタデータシートの作成
            ws_metadata = wb.create_sheet("メタデータ")
            
            # 処理統計
            total_images = len(results)
            successful_images = sum(1 for r in results if r.get('success'))
            failed_images = total_images - successful_images
            total_processing_time = sum(r.get('processing_time', 0) for r in results)
            
            metadata = [
                ['項目', '値'],
                ['総画像数', total_images],
                ['成功数', successful_images],
                ['失敗数', failed_images],
                ['成功率', f"{successful_images/total_images*100:.1f}%" if total_images > 0 else "0%"],
                ['総処理時間', f"{total_processing_time:.2f}秒"],
                ['平均処理時間', f"{total_processing_time/total_images:.2f}秒" if total_images > 0 else "0秒"]
            ]
            
            for row_idx, row_data in enumerate(metadata, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_metadata.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 列幅の調整
            ws_metadata.column_dimensions['A'].width = 15
            ws_metadata.column_dimensions['B'].width = 20
            
            # ファイルの保存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Excelエクスポートに失敗しました: {str(e)}")
            return None
    
    def export_to_pandas_dataframe(self, results: List[Dict[str, Any]]) -> Optional['pd.DataFrame']:
        """
        結果をPandas DataFrameに変換
        
        Args:
            results: OCR結果のリスト
            
        Returns:
            Pandas DataFrame
        """
        if not PANDAS_AVAILABLE:
            return None
        
        if not results:
            return None
        
        try:
            # データの準備
            data = []
            for i, result in enumerate(results):
                data.append({
                    '画像番号': f"画像 {i+1}",
                    'ファイル名': result.get('filename', f'image_{i+1}'),
                    '抽出テキスト': result.get('text', ''),
                    '言語': result.get('language', 'unknown'),
                    '処理時間(秒)': round(result.get('processing_time', 0), 2),
                    'ステータス': '成功' if result.get('success') else '失敗',
                    'エラー': result.get('error', '')
                })
            
            return pd.DataFrame(data)
            
        except Exception as e:
            st.error(f"DataFrame変換に失敗しました: {str(e)}")
            return None
    
    def create_summary_report(self, results: List[Dict[str, Any]]) -> str:
        """
        結果のサマリーレポートを作成
        
        Args:
            results: OCR結果のリスト
            
        Returns:
            サマリーレポートのテキスト
        """
        if not results:
            return "処理結果がありません"
        
        total_images = len(results)
        successful_images = sum(1 for r in results if r.get('success'))
        failed_images = total_images - successful_images
        total_processing_time = sum(r.get('processing_time', 0) for r in results)
        total_characters = sum(len(r.get('text', '')) for r in results)
        
        # 言語統計
        language_stats = {}
        for result in results:
            if result.get('success'):
                lang = result.get('language', 'unknown')
                language_stats[lang] = language_stats.get(lang, 0) + 1
        
        report = f"""
OCR処理結果サマリー
{'=' * 50}

基本統計:
- 総画像数: {total_images}
- 成功数: {successful_images}
- 失敗数: {failed_images}
- 成功率: {successful_images/total_images*100:.1f}% ({successful_images}/{total_images})

処理時間:
- 総処理時間: {total_processing_time:.2f}秒
- 平均処理時間: {total_processing_time/total_images:.2f}秒/画像

テキスト統計:
- 総文字数: {total_characters:,}文字
- 平均文字数: {total_characters/successful_images:.0f}文字/画像 (成功した画像のみ)

検出言語:
"""
        
        for lang, count in sorted(language_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = count / successful_images * 100
            report += f"- {lang}: {count}件 ({percentage:.1f}%)\n"
        
        if failed_images > 0:
            report += f"\n失敗した画像:\n"
            for i, result in enumerate(results):
                if not result.get('success'):
                    report += f"- 画像 {i+1}: {result.get('error', '不明なエラー')}\n"
        
        return report


# グローバルなエクスポートマネージャーインスタンス
export_manager = ExportManager()
